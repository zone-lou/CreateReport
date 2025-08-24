from openpyxl import load_workbook

address={
    "控制点_K1_已知X":0.000,
    "控制点_K1_已知Y":0.000,
    "控制点_K1_检测X":0.000,
    "控制点_K1_检测Y":0.000,
    "控制点_K1_差值":0.000,
    "控制点_K2_已知X":0.000,
    "控制点_K2_已知Y":0.000,
    "控制点_K2_检测X":0.000,
    "控制点_K2_检测Y":0.000,
    "控制点_K2_差值":0.000,
    "控制点_K3_已知X":0.000,
    "控制点_K3_已知Y":0.000,
    "控制点_K3_检测X":0.000,
    "控制点_K3_检测Y":0.000,
    "控制点_K3_差值":0.000,
    "控制点_K1K2_反算边长":0.000,
    "控制点_K1K2_检测边长":0.000,
    "控制点_K1K2_边长差值":0.000,
    "控制点_K1K2_相对误差":0.000,
    "控制点_K2K3_反算边长":0.000,
    "控制点_K2K3_检测边长":0.000,
    "控制点_K2K3_边长差值":0.000,
    "控制点_K2K3_相对误差":0.000,
    "界址点_J1_已知X":0.000,
    "界址点_J1_已知Y":0.000,
    "界址点_J1_检测X":0.000,
    "界址点_J1_检测Y":0.000,
    "界址点_J1_差值":0.000,
    "界址点_J2_已知X":0.000,
    "界址点_J2_已知Y":0.000,
    "界址点_J2_检测X":0.000,
    "界址点_J2_检测Y":0.000,
    "界址点_J2_差值":0.000,
    "界址点_J3_已知X":0.000,
    "界址点_J3_已知Y":0.000,
    "界址点_J3_检测X":0.000,
    "界址点_J3_检测Y":0.000,
    "界址点_J3_差值":0.000,
    "界址点_J4_已知X":0.000,
    "界址点_J4_已知Y":0.000,
    "界址点_J4_检测X":0.000,
    "界址点_J4_检测Y":0.000,
    "界址点_J4_差值":0.000,
    "界址点_J1J2_反算边长":0.000,
    "界址点_J1J2_检测边长":0.000,
    "界址点_J1J2_边长差值":0.000,
    "界址点_J2J3_反算边长":0.000,
    "界址点_J2J3__检测边长":0.000,
    "界址点_J2J3__边长差值":0.000,
    "界址点_J3J4_反算边长":0.000,
    "界址点_J3J4_检测边长":0.000,
    "界址点_J3J4_边长差值":0.000,
    "界址点_J4J1_反算边长":0.000,
    "界址点_J4J1_检测边长":0.000,
    "界址点_J4J1_边长差值":0.000,
    "地块面积":0.000
}

wb = load_workbook(r'/Users/louzeyu/Downloads/不动产报告生成/不动产报告/坐标数据.xlsx',data_only=True)


"""
excel 的检测点都是随机后算出来的，所有只需要取到
控制点_K1_已知X 控制点_K1_已知Y
控制点_K2_已知X 控制点_K2_已知Y
控制点_K3_已知X 控制点_K3_已知Y
界址点_J1_已知X 界址点_J1_已知Y
界址点_J2_已知X 界址点_J2_已知Y
界址点_J3_已知X 界址点_J3_已知Y
界址点_J4_已知X 界址点_J4_已知Y
地块面积
"""
ws = wb['Sheet1']
##k1
address["控制点_K1_已知X"] = ws.cell(row=5, column=2).value
address["控制点_K1_已知Y"] = ws.cell(row=5, column=3).value
address["控制点_K1_检测X"] = ws.cell(row=5, column=4).value
address["控制点_K1_检测Y"] = ws.cell(row=5, column=5).value
address["控制点_K1_差值"] = round(ws.cell(row=5, column=6).value,3)
##k2
address["控制点_K2_已知X"] = ws.cell(row=6, column=2).value
address["控制点_K2_已知Y"] = ws.cell(row=6, column=3).value
address["控制点_K2_检测X"] = ws.cell(row=6, column=4).value
address["控制点_K2_检测Y"] = ws.cell(row=6, column=5).value
address["控制点_K2_差值"] = round(float(ws.cell(row=6, column=6).value),3)
##k3
address["控制点_K3_已知X"] = ws.cell(row=7, column=2).value
address["控制点_K3_已知Y"] = ws.cell(row=7, column=3).value
address["控制点_K3_检测X"] = ws.cell(row=7, column=4).value
address["控制点_K3_检测Y"] = ws.cell(row=7, column=5).value
address["控制点_K3_差值"] = round(float(ws.cell(row=7, column=6).value),3)
##k1k2
address["控制点_K1K2_反算边长"] = round(float(ws.cell(row=12, column=6).value),3)
address["控制点_K1K2_检测边长"] = round(float(ws.cell(row=12, column=7).value),3)
address["控制点_K1K2_边长差值"] = round(float(ws.cell(row=12, column=8).value),3)
address["控制点_K1K2_相对误差"] = ws.cell(row=12, column=9).value
##k2k3
address["控制点_K2K3_反算边长"] = round(float(ws.cell(row=13, column=6).value),3)
address["控制点_K2K3_检测边长"] = round(float(ws.cell(row=13, column=7).value),3)
address["控制点_K2K3_边长差值"] = round(float(ws.cell(row=13, column=8).value),3)
address["控制点_K2K3_相对误差"] = ws.cell(row=13, column=9).value
##j1
address["界址点_J1_已知X"] = ws.cell(row=24, column=2).value
address["界址点_J1_已知Y"] = ws.cell(row=24, column=3).value
address["界址点_J1_检测X"] = ws.cell(row=24, column=4).value
address["界址点_J1_检测Y"] = ws.cell(row=24, column=5).value
address["界址点_J1_差值"] = round(float(ws.cell(row=24, column=6).value),3)
##j2
address["界址点_J2_已知X"] = ws.cell(row=25, column=2).value
address["界址点_J2_已知Y"] = ws.cell(row=25, column=3).value
address["界址点_J2_检测X"] = ws.cell(row=25, column=4).value
address["界址点_J2_检测Y"] = ws.cell(row=25, column=5).value
address["界址点_J2_差值"] = round(float(ws.cell(row=25, column=6).value),3)
##j3
address["界址点_J3_已知X"] = ws.cell(row=26, column=2).value
address["界址点_J3_已知Y"] = ws.cell(row=26, column=3).value
address["界址点_J3_检测X"] = ws.cell(row=26, column=4).value
address["界址点_J3_检测Y"] = ws.cell(row=26, column=5).value
address["界址点_J3_差值"] = round(float(ws.cell(row=26, column=6).value),3)
##j4
address["界址点_J4_已知X"] = ws.cell(row=27, column=2).value
address["界址点_J4_已知Y"] = ws.cell(row=27, column=3).value
address["界址点_J4_检测X"] = ws.cell(row=27, column=4).value
address["界址点_J4_检测Y"] = ws.cell(row=27, column=5).value
address["界址点_J4_差值"] = round(float(ws.cell(row=27, column=6).value),3)
##j1j2
address["界址点_J1J2_反算边长"] = round(float(ws.cell(row=32, column=6).value),3)
address["界址点_J1J2_检测边长"] = round(float(ws.cell(row=32, column=7).value),3)
address["界址点_J1J2_边长差值"] = round(float(ws.cell(row=32, column=8).value),3)
##j2j3
address["界址点_J2J3_反算边长"] = round(float(ws.cell(row=33, column=6).value),3)
address["界址点_J2J3_检测边长"] = round(float(ws.cell(row=33, column=7).value),3)
address["界址点_J2J3_边长差值"] = round(float(ws.cell(row=33, column=8).value),3)
##j3j4
address["界址点_J3J4_反算边长"] = round(float(ws.cell(row=34, column=6).value),3)
address["界址点_J3J4_检测边长"] = round(float(ws.cell(row=34, column=7).value),3)
address["界址点_J3J4_边长差值"] = round(float(ws.cell(row=34, column=8).value),3)
##j4j1
address["界址点_J4J1_反算边长"] = round(float(ws.cell(row=35, column=6).value),3)
address["界址点_J4J1_检测边长"] = round(float(ws.cell(row=35, column=7).value),3)
address["界址点_J4J1_边长差值"] = round(float(ws.cell(row=35, column=8).value),3)
##地块面积
address["地块面积"] = ws.cell(row=38, column=4).value
print(address)